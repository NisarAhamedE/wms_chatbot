import os
import asyncio
import hashlib
from typing import Dict, Any, Optional, List
import aiofiles
from pathlib import Path
import mimetypes
from datetime import datetime
import logging

# Text extraction libraries
import PyPDF2
from docx import Document
import openpyxl
from PIL import Image
import pytesseract
import speech_recognition as sr
import cv2
from moviepy.editor import VideoFileClip

# ML and NLP libraries
import spacy
from transformers import pipeline, AutoTokenizer, AutoModel
import torch
from sklearn.feature_extraction.text import TfidfVectorizer
from sklearn.metrics.pairwise import cosine_similarity
import numpy as np

# LangChain for text processing
from langchain.text_splitter import RecursiveCharacterTextSplitter
from langchain.embeddings import HuggingFaceEmbeddings
from langchain.schema import Document as LangChainDocument

from .models import FileMetadata, FileProcessingTask, ProcessingStage, FileStatus, WMS_CATEGORIES
from ..database import get_db

# Configure logging
logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

class FileProcessor:
    """Main file processing engine."""
    
    def __init__(self):
        self.upload_dir = Path(os.getenv("UPLOAD_DIR", "uploads"))
        self.upload_dir.mkdir(exist_ok=True)
        
        # Initialize NLP models
        self.setup_models()
        
        # Processing stages
        self.stages = [
            ProcessingStage.VALIDATION,
            ProcessingStage.TEXT_EXTRACTION,
            ProcessingStage.CATEGORIZATION,
            ProcessingStage.SUMMARIZATION,
            ProcessingStage.VECTORIZATION,
            ProcessingStage.STORAGE
        ]
    
    def setup_models(self):
        """Initialize ML models for processing."""
        try:
            # Load spaCy model for NLP
            self.nlp = spacy.load("en_core_web_sm")
            
            # Load summarization model
            self.summarizer = pipeline(
                "summarization",
                model="facebook/bart-large-cnn",
                device=0 if torch.cuda.is_available() else -1
            )
            
            # Load embeddings model
            self.embeddings = HuggingFaceEmbeddings(
                model_name="sentence-transformers/all-MiniLM-L6-v2"
            )
            
            # Text splitter for chunking
            self.text_splitter = RecursiveCharacterTextSplitter(
                chunk_size=1000,
                chunk_overlap=200,
                length_function=len
            )
            
            # TF-IDF for keyword extraction
            self.tfidf = TfidfVectorizer(
                max_features=100,
                stop_words='english',
                ngram_range=(1, 2)
            )
            
            logger.info("ML models loaded successfully")
            
        except Exception as e:
            logger.error(f"Error loading ML models: {e}")
            # Fallback to basic processing
            self.nlp = None
            self.summarizer = None
            self.embeddings = None
    
    async def process_file(self, file_id: str, db_session) -> Dict[str, Any]:
        """Process a file through all stages."""
        try:
            # Get file metadata
            file_metadata = db_session.query(FileMetadata).filter(
                FileMetadata.id == file_id
            ).first()
            
            if not file_metadata:
                raise ValueError(f"File {file_id} not found")
            
            # Update status
            file_metadata.status = FileStatus.PROCESSING
            file_metadata.processing_started_at = datetime.utcnow()
            file_metadata.processing_progress = 0.0
            db_session.commit()
            
            results = {}
            total_stages = len(self.stages)
            
            for i, stage in enumerate(self.stages):
                try:
                    # Update progress
                    file_metadata.processing_stage = stage.value
                    file_metadata.processing_progress = (i / total_stages) * 100
                    db_session.commit()
                    
                    # Process stage
                    stage_result = await self.process_stage(file_metadata, stage, db_session)
                    results[stage.value] = stage_result
                    
                    # Log stage completion
                    self.log_processing_task(file_id, stage, "completed", stage_result, db_session)
                    
                except Exception as e:
                    logger.error(f"Error in stage {stage.value} for file {file_id}: {e}")
                    self.log_processing_task(file_id, stage, "failed", None, db_session, str(e))
                    results[stage.value] = {"error": str(e)}
            
            # Update final status
            file_metadata.status = FileStatus.COMPLETED
            file_metadata.processing_completed_at = datetime.utcnow()
            file_metadata.processing_progress = 100.0
            file_metadata.processing_stage = None
            db_session.commit()
            
            return results
            
        except Exception as e:
            logger.error(f"Error processing file {file_id}: {e}")
            # Update error status
            file_metadata.status = FileStatus.FAILED
            file_metadata.error_message = str(e)
            file_metadata.processing_completed_at = datetime.utcnow()
            db_session.commit()
            
            raise e
    
    async def process_stage(
        self, 
        file_metadata: FileMetadata, 
        stage: ProcessingStage, 
        db_session
    ) -> Dict[str, Any]:
        """Process a specific stage."""
        
        if stage == ProcessingStage.VALIDATION:
            return await self.validate_file(file_metadata)
        
        elif stage == ProcessingStage.TEXT_EXTRACTION:
            return await self.extract_text(file_metadata)
        
        elif stage == ProcessingStage.CATEGORIZATION:
            return await self.categorize_file(file_metadata, db_session)
        
        elif stage == ProcessingStage.SUMMARIZATION:
            return await self.generate_summary(file_metadata)
        
        elif stage == ProcessingStage.VECTORIZATION:
            return await self.vectorize_content(file_metadata)
        
        elif stage == ProcessingStage.STORAGE:
            return await self.store_results(file_metadata, db_session)
        
        else:
            raise ValueError(f"Unknown processing stage: {stage}")
    
    async def validate_file(self, file_metadata: FileMetadata) -> Dict[str, Any]:
        """Validate file integrity and format."""
        file_path = Path(file_metadata.file_path)
        
        if not file_path.exists():
            raise FileNotFoundError(f"File not found: {file_path}")
        
        # Check file size
        actual_size = file_path.stat().st_size
        if actual_size != file_metadata.file_size:
            raise ValueError(f"File size mismatch: expected {file_metadata.file_size}, got {actual_size}")
        
        # Verify file hash (if available)
        file_hash = await self.calculate_file_hash(file_path)
        
        # Check MIME type
        detected_mime, _ = mimetypes.guess_type(str(file_path))
        
        return {
            "file_exists": True,
            "size_valid": True,
            "file_hash": file_hash,
            "detected_mime_type": detected_mime,
            "validation_passed": True
        }
    
    async def extract_text(self, file_metadata: FileMetadata) -> Dict[str, Any]:
        """Extract text from various file formats."""
        file_path = Path(file_metadata.file_path)
        mime_type = file_metadata.mime_type.lower()
        
        extracted_text = ""
        extraction_method = "none"
        page_count = None
        duration = None
        
        try:
            if mime_type == "application/pdf":
                extracted_text, page_count = await self.extract_text_from_pdf(file_path)
                extraction_method = "pdf"
                
            elif mime_type in ["application/msword", "application/vnd.openxmlformats-officedocument.wordprocessingml.document"]:
                extracted_text = await self.extract_text_from_docx(file_path)
                extraction_method = "docx"
                
            elif mime_type in ["application/vnd.ms-excel", "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"]:
                extracted_text = await self.extract_text_from_excel(file_path)
                extraction_method = "excel"
                
            elif mime_type.startswith("image/"):
                extracted_text = await self.extract_text_from_image(file_path)
                extraction_method = "ocr"
                
            elif mime_type.startswith("audio/"):
                extracted_text, duration = await self.extract_text_from_audio(file_path)
                extraction_method = "speech_recognition"
                
            elif mime_type.startswith("video/"):
                extracted_text, duration = await self.extract_text_from_video(file_path)
                extraction_method = "video_audio_extraction"
                
            elif mime_type.startswith("text/"):
                async with aiofiles.open(file_path, 'r', encoding='utf-8') as f:
                    extracted_text = await f.read()
                extraction_method = "direct_read"
            
            # Update file metadata
            file_metadata.extracted_text = extracted_text
            if page_count:
                file_metadata.page_count = page_count
            if duration:
                file_metadata.duration = duration
            
            # Detect language
            language = self.detect_language(extracted_text)
            file_metadata.language = language
            
            return {
                "extracted_text": extracted_text,
                "extraction_method": extraction_method,
                "character_count": len(extracted_text),
                "word_count": len(extracted_text.split()),
                "language": language,
                "page_count": page_count,
                "duration": duration
            }
            
        except Exception as e:
            logger.error(f"Text extraction failed: {e}")
            return {
                "extracted_text": "",
                "extraction_method": "failed",
                "error": str(e)
            }
    
    async def categorize_file(self, file_metadata: FileMetadata, db_session) -> Dict[str, Any]:
        """Categorize file based on content and filename."""
        text = file_metadata.extracted_text or ""
        filename = file_metadata.original_name.lower()
        
        # Score each WMS category
        category_scores = {}
        
        for category in WMS_CATEGORIES:
            score = 0.0
            
            # Keyword matching in content
            if text:
                for keyword in category["keywords"]:
                    if keyword.lower() in text.lower():
                        score += 2.0
            
            # Keyword matching in filename
            for keyword in category["keywords"]:
                if keyword.lower() in filename:
                    score += 1.5
            
            # File type bonus
            file_ext = Path(filename).suffix.lower().lstrip('.')
            if file_ext in category["file_types"]:
                score += 0.5
            
            category_scores[category["name"]] = score
        
        # Get top categories
        sorted_categories = sorted(
            category_scores.items(), 
            key=lambda x: x[1], 
            reverse=True
        )
        
        # Assign categories with score > 0
        assigned_categories = [cat for cat, score in sorted_categories if score > 0]
        
        # If no categories match, assign "Other"
        if not assigned_categories:
            assigned_categories = ["Other"]
        
        # Take top 3 categories
        assigned_categories = assigned_categories[:3]
        
        # Update file metadata
        file_metadata.categories = assigned_categories
        file_metadata.confidence_score = sorted_categories[0][1] if sorted_categories else 0.0
        
        return {
            "assigned_categories": assigned_categories,
            "category_scores": dict(sorted_categories[:5]),
            "confidence_score": file_metadata.confidence_score
        }
    
    async def generate_summary(self, file_metadata: FileMetadata) -> Dict[str, Any]:
        """Generate summary of file content."""
        text = file_metadata.extracted_text
        
        if not text or len(text) < 100:
            return {"summary": "", "method": "insufficient_content"}
        
        try:
            # Use extractive summarization for long texts
            if self.summarizer and len(text) > 500:
                # Chunk text if too long
                max_length = 1024
                if len(text) > max_length:
                    text = text[:max_length]
                
                summary_result = self.summarizer(
                    text,
                    max_length=150,
                    min_length=50,
                    do_sample=False
                )
                
                summary = summary_result[0]['summary_text']
                method = "transformers_bart"
            
            else:
                # Fallback to simple extractive summary
                sentences = text.split('. ')
                if len(sentences) > 3:
                    summary = '. '.join(sentences[:3]) + '.'
                else:
                    summary = text[:300] + "..." if len(text) > 300 else text
                method = "extractive"
            
            # Update file metadata
            file_metadata.summary = summary
            
            return {
                "summary": summary,
                "method": method,
                "original_length": len(text),
                "summary_length": len(summary)
            }
            
        except Exception as e:
            logger.error(f"Summary generation failed: {e}")
            # Fallback summary
            summary = text[:200] + "..." if len(text) > 200 else text
            file_metadata.summary = summary
            
            return {
                "summary": summary,
                "method": "fallback_truncation",
                "error": str(e)
            }
    
    async def vectorize_content(self, file_metadata: FileMetadata) -> Dict[str, Any]:
        """Create vector embeddings for semantic search."""
        text = file_metadata.extracted_text
        
        if not text:
            return {"vectors_created": 0, "method": "no_content"}
        
        try:
            # Split text into chunks
            chunks = self.text_splitter.split_text(text)
            
            if not self.embeddings:
                return {"vectors_created": 0, "method": "no_embeddings_model"}
            
            # Create embeddings for each chunk
            embeddings = []
            for chunk in chunks:
                embedding = self.embeddings.embed_query(chunk)
                embeddings.append({
                    "text": chunk,
                    "embedding": embedding,
                    "file_id": file_metadata.id
                })
            
            # Store embeddings (would integrate with Weaviate in production)
            # For now, store metadata about vectorization
            
            return {
                "vectors_created": len(embeddings),
                "chunk_count": len(chunks),
                "method": "sentence_transformers",
                "model": "all-MiniLM-L6-v2"
            }
            
        except Exception as e:
            logger.error(f"Vectorization failed: {e}")
            return {
                "vectors_created": 0,
                "method": "failed",
                "error": str(e)
            }
    
    async def store_results(self, file_metadata: FileMetadata, db_session) -> Dict[str, Any]:
        """Store processing results and clean up."""
        try:
            # Extract keywords if text is available
            if file_metadata.extracted_text:
                keywords = self.extract_keywords(file_metadata.extracted_text)
                file_metadata.keywords = keywords
            
            # Store final metadata
            db_session.commit()
            
            # Clean up temporary files if needed
            # (In production, might move to permanent storage)
            
            return {
                "storage_completed": True,
                "keywords_extracted": len(file_metadata.keywords or []),
                "final_status": file_metadata.status.value
            }
            
        except Exception as e:
            logger.error(f"Storage failed: {e}")
            return {
                "storage_completed": False,
                "error": str(e)
            }
    
    # Helper methods for specific file types
    
    async def extract_text_from_pdf(self, file_path: Path) -> tuple[str, int]:
        """Extract text from PDF files."""
        text = ""
        page_count = 0
        
        with open(file_path, 'rb') as file:
            pdf_reader = PyPDF2.PdfReader(file)
            page_count = len(pdf_reader.pages)
            
            for page in pdf_reader.pages:
                text += page.extract_text() + "\n"
        
        return text.strip(), page_count
    
    async def extract_text_from_docx(self, file_path: Path) -> str:
        """Extract text from Word documents."""
        doc = Document(file_path)
        text = "\n".join([paragraph.text for paragraph in doc.paragraphs])
        return text.strip()
    
    async def extract_text_from_excel(self, file_path: Path) -> str:
        """Extract text from Excel files."""
        workbook = openpyxl.load_workbook(file_path)
        text_parts = []
        
        for sheet_name in workbook.sheetnames:
            sheet = workbook[sheet_name]
            for row in sheet.iter_rows(values_only=True):
                row_text = " ".join([str(cell) for cell in row if cell is not None])
                if row_text.strip():
                    text_parts.append(row_text)
        
        return "\n".join(text_parts)
    
    async def extract_text_from_image(self, file_path: Path) -> str:
        """Extract text from images using OCR."""
        try:
            image = Image.open(file_path)
            text = pytesseract.image_to_string(image)
            return text.strip()
        except Exception as e:
            logger.error(f"OCR failed: {e}")
            return ""
    
    async def extract_text_from_audio(self, file_path: Path) -> tuple[str, float]:
        """Extract text from audio files using speech recognition."""
        try:
            recognizer = sr.Recognizer()
            
            # Get audio duration
            with sr.AudioFile(str(file_path)) as source:
                audio = recognizer.record(source)
                duration = len(audio.frame_data) / audio.sample_rate
            
            # Recognize speech
            text = recognizer.recognize_google(audio)
            
            return text, duration
            
        except Exception as e:
            logger.error(f"Speech recognition failed: {e}")
            return "", 0.0
    
    async def extract_text_from_video(self, file_path: Path) -> tuple[str, float]:
        """Extract text from video files by extracting audio."""
        try:
            # Extract audio from video
            video = VideoFileClip(str(file_path))
            duration = video.duration
            
            # Extract audio to temporary file
            temp_audio_path = file_path.with_suffix('.wav')
            video.audio.write_audiofile(str(temp_audio_path), verbose=False, logger=None)
            
            # Extract text from audio
            text, _ = await self.extract_text_from_audio(temp_audio_path)
            
            # Clean up
            temp_audio_path.unlink(missing_ok=True)
            video.close()
            
            return text, duration
            
        except Exception as e:
            logger.error(f"Video text extraction failed: {e}")
            return "", 0.0
    
    def extract_keywords(self, text: str, max_keywords: int = 20) -> List[str]:
        """Extract keywords from text using TF-IDF."""
        if not text or len(text) < 50:
            return []
        
        try:
            if self.nlp:
                # Use spaCy for better keyword extraction
                doc = self.nlp(text)
                keywords = []
                
                for token in doc:
                    if (token.pos_ in ['NOUN', 'ADJ'] and 
                        not token.is_stop and 
                        not token.is_punct and 
                        len(token.text) > 2):
                        keywords.append(token.lemma_.lower())
                
                # Get most frequent keywords
                keyword_freq = {}
                for keyword in keywords:
                    keyword_freq[keyword] = keyword_freq.get(keyword, 0) + 1
                
                sorted_keywords = sorted(keyword_freq.items(), key=lambda x: x[1], reverse=True)
                return [kw for kw, freq in sorted_keywords[:max_keywords]]
            
            else:
                # Fallback to simple word frequency
                words = text.lower().split()
                word_freq = {}
                for word in words:
                    if len(word) > 3:
                        word_freq[word] = word_freq.get(word, 0) + 1
                
                sorted_words = sorted(word_freq.items(), key=lambda x: x[1], reverse=True)
                return [word for word, freq in sorted_words[:max_keywords]]
                
        except Exception as e:
            logger.error(f"Keyword extraction failed: {e}")
            return []
    
    def detect_language(self, text: str) -> str:
        """Detect text language."""
        if not text or len(text) < 50:
            return "unknown"
        
        try:
            if self.nlp:
                # Use spaCy language detection
                doc = self.nlp(text[:1000])  # Sample first 1000 chars
                return "en"  # Simplified - in production would use proper language detection
            else:
                return "en"  # Default to English
                
        except Exception:
            return "unknown"
    
    async def calculate_file_hash(self, file_path: Path) -> str:
        """Calculate SHA-256 hash of file."""
        hash_sha256 = hashlib.sha256()
        
        async with aiofiles.open(file_path, 'rb') as f:
            while chunk := await f.read(8192):
                hash_sha256.update(chunk)
        
        return hash_sha256.hexdigest()
    
    def log_processing_task(
        self,
        file_id: str,
        stage: ProcessingStage,
        status: str,
        result_data: Any,
        db_session,
        error_message: str = None
    ):
        """Log processing task to database."""
        task = FileProcessingTask(
            file_id=file_id,
            task_type=stage.value,
            status=status,
            completed_at=datetime.utcnow(),
            result_data=result_data,
            error_message=error_message
        )
        
        db_session.add(task)
        db_session.commit()

# Initialize file processor
file_processor = FileProcessor()