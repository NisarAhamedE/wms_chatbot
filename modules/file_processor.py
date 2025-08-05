import os
import time
import mimetypes
from pathlib import Path
from typing import Dict, List, Optional, Any, Tuple
import threading
import queue
import asyncio

# Document processing libraries
import PyPDF2
from docx import Document
import openpyxl
from openpyxl import load_workbook
import pandas as pd
import markdown
from bs4 import BeautifulSoup

# Image processing
from PIL import Image
import cv2
import numpy as np
import pytesseract

# Azure OpenAI for advanced text extraction
import openai
from openai import AzureOpenAI

from .logger import LoggerMixin
from .config_manager import ConfigManager

class FileProcessor(LoggerMixin):
    """
    Handles multi-format document processing with OCR and AI capabilities
    """
    
    def __init__(self, db_manager, config_manager: ConfigManager = None):
        super().__init__()
        
        self.db_manager = db_manager
        self.config_manager = config_manager or ConfigManager()
        
        # Initialize supported formats
        self.supported_formats = {
            'documents': ['.txt', '.doc', '.docx', '.pdf', '.md', '.rtf'],
            'spreadsheets': ['.xlsx', '.xls', '.csv'],
            'images': ['.png', '.jpg', '.jpeg', '.gif', '.bmp', '.tiff'],
            'presentations': ['.ppt', '.pptx'],
            'web': ['.html', '.htm']
        }
        
        # Initialize processors
        self.processors = {
            '.pdf': PDFProcessor(),
            '.doc': DocProcessor(),
            '.docx': DocxProcessor(),
            '.xlsx': ExcelProcessor(),
            '.xls': ExcelProcessor(),
            '.csv': CSVProcessor(),
            '.txt': TextProcessor(),
            '.md': MarkdownProcessor(),
            '.html': HTMLProcessor(),
            '.png': ImageProcessor(),
            '.jpg': ImageProcessor(),
            '.jpeg': ImageProcessor(),
            '.gif': ImageProcessor(),
            '.bmp': ImageProcessor(),
            '.tiff': ImageProcessor()
        }
        
        # Initialize Azure OpenAI client
        self.azure_client = None
        self.init_azure_client()
        
        # Processing queue
        self.processing_queue = queue.Queue()
        self.processing_thread = None
        self.start_processing_thread()
        
        self.log_info("File processor initialized")
    
    def init_azure_client(self):
        """Initialize Azure OpenAI client"""
        try:
            azure_config = self.config_manager.get_azure_config()
            if self.config_manager.is_azure_configured():
                self.azure_client = AzureOpenAI(
                    api_key=azure_config["api_key"],
                    api_version=azure_config["api_version"],
                    azure_endpoint=azure_config["endpoint"]
                )
                self.log_info("Azure OpenAI client initialized")
            else:
                self.log_warning("Azure OpenAI not configured - using fallback OCR")
        except Exception as e:
            self.log_error(f"Failed to initialize Azure OpenAI: {e}")
    
    def start_processing_thread(self):
        """Start background processing thread"""
        self.processing_thread = threading.Thread(target=self._process_queue, daemon=True)
        self.processing_thread.start()
    
    def _process_queue(self):
        """Process items in the background queue"""
        while True:
            try:
                item = self.processing_queue.get(timeout=1)
                if item is None:
                    break
                
                self._handle_queue_item(item)
                
            except queue.Empty:
                continue
            except Exception as e:
                self.log_error(f"Error processing queue item: {e}")
    
    def _handle_queue_item(self, item):
        """Handle individual queue items"""
        try:
            item_type = item.get('type')
            callback = item.get('callback')
            
            if item_type == 'file':
                result = self.process_file(item['file_path'])
            elif item_type == 'text':
                result = self.process_text(item['text'], item.get('metadata', {}))
            elif item_type == 'screenshot':
                result = self.process_screenshot(item['screenshot_path'])
            else:
                self.log_error(f"Unknown item type: {item_type}")
                return
            
            # Call callback if provided
            if callback:
                callback(result)
                
        except Exception as e:
            self.log_error(f"Error handling queue item: {e}")
    
    def process_file(self, file_path: str) -> Dict[str, Any]:
        """
        Process a file and extract text content
        
        Args:
            file_path: Path to the file
            
        Returns:
            Dictionary with processing results
        """
        try:
            file_path = Path(file_path)
            
            # Validate file
            if not file_path.exists():
                raise FileNotFoundError(f"File not found: {file_path}")
            
            # Check file size
            file_size = file_path.stat().st_size
            max_size = self.config_manager.get("file_processing.max_file_size", 50 * 1024 * 1024)
            if file_size > max_size:
                raise ValueError(f"File too large: {file_size} bytes (max: {max_size})")
            
            # Get file extension
            file_ext = file_path.suffix.lower()
            
            # Check if format is supported
            if file_ext not in self.processors:
                raise ValueError(f"Unsupported file format: {file_ext}")
            
            # Process file
            start_time = time.time()
            processor = self.processors[file_ext]
            content = processor.extract_text(str(file_path))
            processing_time = time.time() - start_time
            
            # Prepare metadata
            metadata = {
                'file_size': file_size,
                'processing_time': processing_time,
                'processor': processor.__class__.__name__,
                'file_extension': file_ext,
                'mime_type': mimetypes.guess_type(str(file_path))[0]
            }
            
            # Store in database
            document_id = self.db_manager.store_document(str(file_path), content, metadata)
            
            result = {
                'success': True,
                'document_id': document_id,
                'content': content,
                'metadata': metadata,
                'file_path': str(file_path)
            }
            
            self.log_info(f"File processed successfully: {file_path.name}")
            return result
            
        except Exception as e:
            self.log_error(f"Error processing file {file_path}: {e}")
            return {
                'success': False,
                'error': str(e),
                'file_path': str(file_path)
            }
    
    def process_text(self, text: str, metadata: Dict[str, Any] = None) -> Dict[str, Any]:
        """
        Process text input directly
        
        Args:
            text: Text content
            metadata: Additional metadata
            
        Returns:
            Dictionary with processing results
        """
        try:
            if not text.strip():
                raise ValueError("Empty text content")
            
            # Prepare metadata
            if metadata is None:
                metadata = {}
            
            metadata.update({
                'input_type': 'text',
                'text_length': len(text),
                'word_count': len(text.split()),
                'timestamp': time.time()
            })
            
            # Store in database
            document_id = self.db_manager.store_document("text_input", text, metadata)
            
            result = {
                'success': True,
                'document_id': document_id,
                'content': text,
                'metadata': metadata
            }
            
            self.log_info("Text processed successfully")
            return result
            
        except Exception as e:
            self.log_error(f"Error processing text: {e}")
            return {
                'success': False,
                'error': str(e)
            }
    
    def process_screenshot(self, screenshot_path: str) -> Dict[str, Any]:
        """
        Process screenshot and extract text using OCR
        
        Args:
            screenshot_path: Path to screenshot image
            
        Returns:
            Dictionary with processing results
        """
        try:
            screenshot_path = Path(screenshot_path)
            
            if not screenshot_path.exists():
                raise FileNotFoundError(f"Screenshot not found: {screenshot_path}")
            
            # Use image processor for OCR
            processor = self.processors['.png']  # Use PNG processor for all images
            content = processor.extract_text(str(screenshot_path))
            
            # Prepare metadata
            metadata = {
                'input_type': 'screenshot',
                'image_path': str(screenshot_path),
                'image_size': screenshot_path.stat().st_size,
                'timestamp': time.time()
            }
            
            # Store in database
            document_id = self.db_manager.store_document(str(screenshot_path), content, metadata)
            
            # Store screenshot separately
            screenshot_id = self.db_manager.store_screenshot(
                document_id, str(screenshot_path), content
            )
            
            result = {
                'success': True,
                'document_id': document_id,
                'screenshot_id': screenshot_id,
                'content': content,
                'metadata': metadata,
                'screenshot_path': str(screenshot_path)
            }
            
            self.log_info(f"Screenshot processed successfully: {screenshot_path.name}")
            return result
            
        except Exception as e:
            self.log_error(f"Error processing screenshot {screenshot_path}: {e}")
            return {
                'success': False,
                'error': str(e),
                'screenshot_path': str(screenshot_path)
            }
    
    def add_to_queue(self, item_type: str, **kwargs):
        """Add item to processing queue"""
        item = {'type': item_type, **kwargs}
        self.processing_queue.put(item)
    
    def get_supported_formats(self) -> Dict[str, List[str]]:
        """Get supported file formats"""
        return self.supported_formats
    
    def is_supported_format(self, file_path: str) -> bool:
        """Check if file format is supported"""
        file_ext = Path(file_path).suffix.lower()
        return file_ext in self.processors


class BaseProcessor(LoggerMixin):
    """Base class for all document processors"""
    
    def __init__(self):
        super().__init__()
    
    def extract_text(self, file_path: str) -> str:
        """Extract text from file - to be implemented by subclasses"""
        raise NotImplementedError


class PDFProcessor(BaseProcessor):
    """Process PDF files"""
    
    def extract_text(self, file_path: str) -> str:
        try:
            text = ""
            with open(file_path, 'rb') as file:
                pdf_reader = PyPDF2.PdfReader(file)
                for page_num, page in enumerate(pdf_reader.pages):
                    page_text = page.extract_text()
                    text += f"\n--- Page {page_num + 1} ---\n{page_text}\n"
            
            return text.strip()
        except Exception as e:
            self.log_error(f"Error extracting text from PDF {file_path}: {e}")
            raise


class DocProcessor(BaseProcessor):
    """Process DOC files (requires additional libraries)"""
    
    def extract_text(self, file_path: str) -> str:
        try:
            # For DOC files, we might need additional libraries like python-docx2txt
            # For now, return a placeholder
            return f"DOC file content from {file_path} - requires additional processing"
        except Exception as e:
            self.log_error(f"Error extracting text from DOC {file_path}: {e}")
            raise


class DocxProcessor(BaseProcessor):
    """Process DOCX files"""
    
    def extract_text(self, file_path: str) -> str:
        try:
            doc = Document(file_path)
            text = ""
            for paragraph in doc.paragraphs:
                text += paragraph.text + "\n"
            return text.strip()
        except Exception as e:
            self.log_error(f"Error extracting text from DOCX {file_path}: {e}")
            raise


class ExcelProcessor(BaseProcessor):
    """Process Excel files"""
    
    def extract_text(self, file_path: str) -> str:
        try:
            workbook = load_workbook(file_path, data_only=True)
            text = ""
            
            for sheet_name in workbook.sheetnames:
                sheet = workbook[sheet_name]
                text += f"\n--- Sheet: {sheet_name} ---\n"
                
                for row in sheet.iter_rows(values_only=True):
                    row_text = " | ".join(str(cell) if cell is not None else "" for cell in row)
                    if row_text.strip():
                        text += row_text + "\n"
            
            return text.strip()
        except Exception as e:
            self.log_error(f"Error extracting text from Excel {file_path}: {e}")
            raise


class CSVProcessor(BaseProcessor):
    """Process CSV files"""
    
    def extract_text(self, file_path: str) -> str:
        try:
            df = pd.read_csv(file_path)
            return df.to_string(index=False)
        except Exception as e:
            self.log_error(f"Error extracting text from CSV {file_path}: {e}")
            raise


class TextProcessor(BaseProcessor):
    """Process text files"""
    
    def extract_text(self, file_path: str) -> str:
        try:
            with open(file_path, 'r', encoding='utf-8') as file:
                return file.read()
        except UnicodeDecodeError:
            # Try with different encoding
            try:
                with open(file_path, 'r', encoding='latin-1') as file:
                    return file.read()
            except Exception as e:
                self.log_error(f"Error reading text file {file_path}: {e}")
                raise
        except Exception as e:
            self.log_error(f"Error extracting text from text file {file_path}: {e}")
            raise


class MarkdownProcessor(BaseProcessor):
    """Process Markdown files"""
    
    def extract_text(self, file_path: str) -> str:
        try:
            with open(file_path, 'r', encoding='utf-8') as file:
                md_content = file.read()
                # Convert markdown to plain text
                html = markdown.markdown(md_content)
                soup = BeautifulSoup(html, 'html.parser')
                return soup.get_text()
        except Exception as e:
            self.log_error(f"Error extracting text from Markdown {file_path}: {e}")
            raise


class HTMLProcessor(BaseProcessor):
    """Process HTML files"""
    
    def extract_text(self, file_path: str) -> str:
        try:
            with open(file_path, 'r', encoding='utf-8') as file:
                html_content = file.read()
                soup = BeautifulSoup(html_content, 'html.parser')
                return soup.get_text()
        except Exception as e:
            self.log_error(f"Error extracting text from HTML {file_path}: {e}")
            raise


class ImageProcessor(BaseProcessor):
    """Process image files with OCR"""
    
    def __init__(self):
        super().__init__()
        self.azure_client = None  # Will be set by FileProcessor
    
    def extract_text(self, file_path: str) -> str:
        try:
            # Try Azure OpenAI Vision API first
            if self.azure_client:
                return self._extract_with_azure(file_path)
            else:
                # Fallback to Tesseract OCR
                return self._extract_with_tesseract(file_path)
        except Exception as e:
            self.log_error(f"Error extracting text from image {file_path}: {e}")
            raise
    
    def _extract_with_azure(self, file_path: str) -> str:
        """Extract text using Azure OpenAI Vision API"""
        try:
            with open(file_path, 'rb') as image_file:
                response = self.azure_client.chat.completions.create(
                    model="gpt-4-vision-preview",
                    messages=[
                        {
                            "role": "user",
                            "content": [
                                {
                                    "type": "text",
                                    "text": "Extract all text from this image. Return only the text content, no explanations."
                                },
                                {
                                    "type": "image_url",
                                    "image_url": {
                                        "url": f"data:image/jpeg;base64,{image_file.read()}"
                                    }
                                }
                            ]
                        }
                    ],
                    max_tokens=4000
                )
                return response.choices[0].message.content
        except Exception as e:
            self.log_error(f"Azure Vision API failed, falling back to Tesseract: {e}")
            return self._extract_with_tesseract(file_path)
    
    def _extract_with_tesseract(self, file_path: str) -> str:
        """Extract text using Tesseract OCR"""
        try:
            # Preprocess image for better OCR
            image = cv2.imread(file_path)
            if image is None:
                raise ValueError(f"Could not load image: {file_path}")
            
            # Convert to grayscale
            gray = cv2.cvtColor(image, cv2.COLOR_BGR2GRAY)
            
            # Apply preprocessing
            # 1. Resize if too small
            height, width = gray.shape
            if width < 100 or height < 100:
                scale_factor = max(100/width, 100/height)
                new_width = int(width * scale_factor)
                new_height = int(height * scale_factor)
                gray = cv2.resize(gray, (new_width, new_height))
            
            # 2. Apply thresholding
            _, thresh = cv2.threshold(gray, 0, 255, cv2.THRESH_BINARY + cv2.THRESH_OTSU)
            
            # 3. Apply morphological operations
            kernel = cv2.getStructuringElement(cv2.MORPH_RECT, (3, 3))
            processed = cv2.morphologyEx(thresh, cv2.MORPH_CLOSE, kernel)
            
            # Extract text using Tesseract
            text = pytesseract.image_to_string(processed, config='--psm 6')
            
            return text.strip()
            
        except Exception as e:
            self.log_error(f"Tesseract OCR failed: {e}")
            raise 